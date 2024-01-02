import os
import pandas as pd
import LoginPage
from Init import get_driver
from VersionSelection import VersionSelection
from LoginPage import LoginPage
import Config
import logging
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import selenium.common.exceptions
from time import sleep
import time
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font


class bluetooth_pairing_test:
    def __init__(self):
        logging.basicConfig(filename='蓝牙配网记录日志.log', filemode='a', level=logging.INFO,
                            format='%(asctime)s - %(name)s - %(levelname)s - %(lineno)d - %(message)s - %(exc_info)s')
        language = 'Chinese'
        versions = 'Chinese'
        agreement_verifycode = (True, True)
        getCode_fillCode = (True, None)
        login_type = "phone"
        self.driver = get_driver()
        version = VersionSelection(self.driver)
        version.version_selection(version=versions, language=language)
        Config.Config.phone = "15250996938"
        login = LoginPage(self.driver)
        self.logger = logging.getLogger(__name__)
        if agreement_verifycode[0]:
            login.login(getCode_fillCode, login_type, language, agreement_verifycode, )
        elif not agreement_verifycode[0]:
            login.login(login_type, language)
        else:
            print("Agreement 变量错误！")

    def screenshot_diy(self, name):
        save_directory = "fail_screenshot"
        os.makedirs(save_directory, exist_ok=True)
        # 构造文件路径
        file_path = os.path.join(save_directory, name)
        screenshot_data = self.driver.get_screenshot_as_png()
        with open(file_path, 'wb') as f:
            f.write(screenshot_data)

    def test_result_statistics(self, file_name):
        # 初始化计数器和最后一次配网次数
        fail_count = 0
        last_attempt = 0
        total_succecc_rate_no = None
        # 尝试打开文件
        try:
            with open(file_name, 'r', encoding='utf-8') as file:
                # 逐行读取文件内容
                lines = file.readlines()
                fail_count = 0

                # 遍历每一行
                for line_number, line in enumerate(lines, start=1):
                    # 判断是否包含连接失败的关键词
                    if "连接失败" in line:
                        fail_count += 1
                    # 获取最后一次配网次数
                    if line.startswith("第") and "次配网" in line:
                        last_attempt = int(line.split("次")[0][1:])
                    if "总的成功率" in line:
                        total_succecc_rate_no = line_number
        except FileNotFoundError:
            print(f"找不到文件：{file_name}")
            self.logger.error(f"找不到文件：{file_name}")
        except Exception as e:
            print(f"发生错误：{e}")
            self.logger.error(f"发生错误：{e}")

        # 打印统计结果
        print(f"连接失败的次数：{fail_count}")
        print(f"最后一次配网次数：{last_attempt}")
        return fail_count, last_attempt, total_succecc_rate_no

    def delete_line_by_number(self, file_path, line_number):
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()

            # 删除指定行
            if 1 <= line_number <= len(lines):
                del lines[line_number - 1]

            with open(file_path, 'w', encoding='utf-8') as file:
                file.writelines(lines)

        except FileNotFoundError:
            print(f"找不到文件：{file_path}")
        except Exception as e:
            print(f"发生错误：{e}")

    def excel_reader(self, file_path):
        try:
            df = pd.read_excel(file_path, index_col=0, engine="openpyxl")
            paired_times = 0
            paired_fail = 0
            twice_paired_fail = 0
            # 获取序列号最后一个数字的值
            for i in df.index[::-1]:
                is_number = isinstance(i, (int, float))
                if is_number:
                    paired_times = i
                    self.logger.info(f"最后一次配网次数：{i}")
                    break
            # 遍历所有行，统计一次配网失败以及二次配网结果,注意经过excel文件后x变成变得字符×
            for row in df.itertuples():
                if '\u2B55' in row or 'x' in row:
                    paired_fail += 1
                if 'x' in row:
                    twice_paired_fail += 1
            self.logger.info(
                f"paired_times: {paired_times}, paired_fail: {paired_fail}, twice_paired_fail: {twice_paired_fail}")
            return paired_times, paired_fail, twice_paired_fail
        except FileNotFoundError:
            print(f"找不到文件：{file_path}, 请确认是否第一次运行脚本")
            self.logger.error(f"找不到文件：{file_path}, 请确认是否第一次运行脚本")
            return 0, 0, 0

    def set_adaptive_column_width(self, writer, data_frame, work_sheet_name="配网结果"):
        """
        :param writer:
        :param data_frame:
        :param work_sheet_name:
        :return:
        因为外部传入DataFrame参数，以及该方法需要在with内使用，而with内无法直接访问DataFrame，需要先将DataFrame通过to_excel传入writer
        """
        # 计算表头的字符宽度
        column_widths = data_frame.columns.to_series().apply(lambda x: len(str(x).encode('utf-8'))).values  # 遍历列名的宽度
        # 计算索引列（序号列）的宽度
        index_width = len(str(data_frame.index.name).encode('utf-8'))
        # 计算每列的最大字符宽度
        max_widths = data_frame.astype(str).map(
            lambda x: len(x.encode('utf-8'))).max().values  # map会遍历data_frame所有元素，而astype会以列的方式将data_frame转换为Series
        # 计算整体最大宽度
        widths = [max(x, y) for x, y in zip(column_widths + 2, max_widths + 2)]
        widths.insert(0, index_width)
        # 设置每列的宽度
        worksheet = writer.sheets[work_sheet_name]  # 获取工作表对象
        for i, width in enumerate(widths):
            col_letter = get_column_letter(i + 1)  # 获取列的字母表示
            worksheet.column_dimensions[col_letter].width = width
            # 设置列名（表头）水平和垂直居中
            cell = worksheet[f"{col_letter}1"]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 设置列值（数据）水平和垂直居中
            for row_num in range(2, len(data_frame) + 2):
                cell = worksheet[f"{col_letter}{row_num}"]
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def set_font_color(self, writer, data_frame, column_names, sheet_name="sheet1", color="FFFFFF",
                       colors_condition=False):
        # 设置字体颜色
        font_color = color
        # 获取工作表对象
        worksheet = writer.sheets[sheet_name]
        # 遍历每一列
        for col_to_color in column_names:
            # 获取列的字母表示
            col_letters = get_column_letter(data_frame.columns.get_loc(col_to_color) + 2)
            # 获取列的所有单元格
            column_cells = worksheet[col_letters]
            # 设置字体颜色
            for cell in column_cells:
                if colors_condition and cell.row > 1:
                    # 获取cell的值
                    if cell.value == "\u2713":
                        cell.font = Font(color="00FF00")
                    elif cell.value == "\u2B55":
                        cell.font = Font(color="0000FF")
                    elif cell.value == "×":
                        cell.font = Font(color="FF0000")
                    else:
                        cell.font = Font(color="000000")
                else:
                    if cell.row > 1:  # 排除列名
                        cell.font = Font(color=font_color)

    def device_results_report(self, device_result, second_try_result):
        if device_result == "配网成功":
            device_excel_result = "\u2713"
        elif device_result == "连接失败" and second_try_result == "二次配网成功":
            device_excel_result = "\u2B55"
        elif device_result == "连接失败" and second_try_result == "二次配网失败":
            device_excel_result = "x"
        else:
            device_excel_result = ""
        return device_excel_result

    def excel_remove_rows(self, file_path, condition):
        try:
            df = pd.read_excel(file_path, index_col=0, engine="openpyxl")
            df.index = df.index.astype(str)
            mask = df.index.str.contains(condition)
            df = df[~mask]
            self.logger.info(f"dataframe：{df}")
            with pd.ExcelWriter(file_path) as writer:
                df.to_excel(writer)
        except FileNotFoundError:
            self.logger.info(f"找不到文件：{file_path}, 请确认是否第一次运行脚本")

    def enter_bluetooth_setup_interface(self, driver):
        try:
            add_device = WebDriverWait(driver, 10).until(
                ec.visibility_of_element_located((By.XPATH, '//android.webkit.WebView[@text="pages/index['
                                                            '2]"]/android.view.View[2]/android.widget.TextView[1]'))
            )
            add_device.click()
        except selenium.common.exceptions.TimeoutException:
            self.logger.error("未找到添加设备按钮")
        try:
            product_selector = WebDriverWait(driver, 10).until(
                ec.visibility_of_element_located((By.XPATH, '//android.webkit.WebView[@text="pages/work/addAhost1['
                                                            '4]"]/android.view.View[3]'))
            )
            product_selector.click()
        except selenium.common.exceptions.TimeoutException:
            self.logger.error("未找到产品选择按钮")
        try:
            location_permission = WebDriverWait(driver, 10).until(
                ec.visibility_of_element_located((By.XPATH, '//android.widget.Button[@text="仅在使用中允许"]'))
            )
            location_permission.click()
        except selenium.common.exceptions.TimeoutException:
            self.logger.error("未找到位置授权按钮")
        try:
            location_permission = WebDriverWait(driver, 10).until(
                ec.visibility_of_element_located((By.XPATH, '//android.widget.Button[@text="始终允许"]'))
            )
            location_permission.click()
        except selenium.common.exceptions.TimeoutException:
            self.logger.error("未找到蓝牙授权按钮")

    def detect_network_setup_interface(self):
        global total_time
        driver = self.driver
        wait = WebDriverWait(driver, 5)
        self.enter_bluetooth_setup_interface(driver)
        sleep(5)
        file_name = "配网结果.txt"
        twice_paired_fail_excel = 0
        # 读取配网结果txt文件，获取已配网的次数
        fail_count, count, total_succecc_rate_no = self.test_result_statistics(file_name=file_name)
        # 删除total_succecc_rate_no行，避免重行
        if total_succecc_rate_no is not None:
            self.delete_line_by_number(file_name, total_succecc_rate_no)
        excel_file_name = '配网结果.xlsx'
        if os.path.exists(excel_file_name):
            # 移除总成功率
            self.excel_remove_rows(excel_file_name, "总的成功率")
            # 读取配网结果xlsx文件，获取已配网的次数
            count_excel, fail_count_excel, twice_paired_fail_excel = self.excel_reader(excel_file_name)
        one_time_setup_successful = 0
        circle_times = 2
        remaining_iterations = circle_times - count
        devices_num = 4
        encoding = 'utf-8'
        file_name = file_name
        current_iteration = 0
        current_twice_pairing_fail = 0
        wifi_name = "小王子"
        wifi_passwd = "sf19960408"
        # 设备信息
        devices = [
            # {"id": "Pintura-blt-L000892", "element": '//android.widget.TextView['
            #                                          f'@resource-id="com.ost.pintura:id/tv_name" and '
            #                                          f'@text="Pintura-blt-L000892"]', "result": None, "second_try_result": None, 'pairing_results_excel': []},
            # {"id": "Pintura-blt-Ltest20", "element": '//android.widget.TextView['
            #                                          f'@resource-id="com.ost.pintura:id/tv_name" and '
            #                                          f'@text="Pintura-blt-Ltest20"]', "result": None, "second_try_result": None, 'pairing_results_excel': []},
            # {"id": "Pintura-blt-L000308", "element": '//android.widget.TextView['
            #                                          f'@resource-id="com.ost.pintura:id/tv_name" and '
            #                                          f'@text="Pintura-blt-L000308"]', "result": None, "second_try_result": None, 'pairing_results_excel': []},
            # {"id": "Pintura-blt-L000329", "element": '//android.widget.TextView['
            #                                          f'@resource-id="com.ost.pintura:id/tv_name" and '
            #                                          f'@text="Pintura-blt-L000329"]', "result": None, "second_try_result": None, 'pairing_results_excel': []},
            {"id": "Pintura-blt-Ltet503", "element": '//android.widget.TextView['
                                                     f'@resource-id="com.ost.pintura:id/tv_name" and '
                                                     f'@text="Pintura-blt-Ltet503"]', "result": None,
             "second_try_result": None, 'pairing_results_excel': []}
        ]
        # 循环配网，直到 circle_times 次
        if remaining_iterations > 0:
            column_names = [device["id"] for device in devices]
            column_names.append("耗时（S）")
            # 创建空列表用于存储每个设备的配网结果
            total_time_list = []
            for i in range(remaining_iterations):
                current_iteration += 1
                devices_successful = 0
                all_devices_found = True
                # 循环前重置result
                for device in devices:
                    device['result'] = None
                    device["second_try_result"] = None
                print(f"第 {current_iteration} 次循环")
                for device in devices:
                    try:
                        element = WebDriverWait(driver, 10).until(
                            ec.visibility_of_element_located((By.XPATH, device["element"]))
                        )
                        element.click()
                        wait.until(ec.staleness_of(element))
                        self.logger.info(f"找到设备：{device['id']}")
                    except selenium.common.exceptions.TimeoutException:
                        self.logger.error(f"未找到设备：{device['id']}")
                        all_devices_found = False
                if not all_devices_found:
                    self.logger.error("存在了设备蓝牙丢失，请打开后输入 continue 继续")
                    while True:
                        user_input = input("请手动打开丢失蓝牙信号设备的蓝牙并输入 continue 继续, 输入 stop 停止脚本")
                        if user_input == "continue":
                            driver.press_keycode(4)
                            is_continue = True
                            try:
                                product_selector = WebDriverWait(driver, 10).until(
                                    ec.visibility_of_element_located((By.XPATH,
                                                                      '//android.webkit.WebView[@text="pages/work/addAhost1[4]"]/android.view.View[3]'))
                                )
                                product_selector.click()
                            except selenium.common.exceptions.TimeoutException:
                                self.logger.error("未找到产品选择按钮")
                            break
                        elif user_input == "stop":
                            is_continue = False
                            break
                        else:
                            print("输入错误，请重新输入")
                    if is_continue:
                        continue
                    else:
                        break

                start_setup_button = WebDriverWait(driver, 10).until(
                    ec.visibility_of_element_located((By.ID, 'com.ost.pintura:id/btn_next'))
                )
                start_setup_button.click()
                # 等待页面加载完成
                if "等待页面加载完成":
                    try:
                        result = WebDriverWait(driver, 15).until(
                            ec.invisibility_of_element_located((By.CLASS_NAME, 'android.widget.ImageView'))
                        )
                    except selenium.common.exceptions.TimeoutException:
                        self.logger.error("未找到页面刷新元素标记")
                        break
                    if result:
                        self.logger.info("页面加载完成")
                        try:
                            wifi_list = WebDriverWait(driver, 10).until(
                                ec.visibility_of_element_located((By.XPATH,
                                                                  '//android.view.ViewGroup[@resource-id="com.ost.pintura:id/swipe_layout"]/android.widget.LinearLayout/android.widget.FrameLayout[1]/android.view.View[2]'))
                            )
                            wifi_list.click()
                            driver.implicitly_wait(1)
                        except selenium.common.exceptions.TimeoutException:
                            self.logger.error("未找到wifi列表")

                if "选择WiFi并开始配网":
                    elments_to_find = {
                        'wifi_list': f'//android.widget.TextView[@text="{wifi_name}"]',
                        'wifi_passwd_element': 'com.ost.pintura:id/et_pwd',
                        'connect_button': 'com.ost.pintura:id/btn_ok'
                    }
                    for element_key, element_value in elments_to_find.items():
                        try:
                            find_result = WebDriverWait(driver, 10).until(
                                ec.visibility_of_element_located(
                                    (By.XPATH, element_value)
                                )
                            )
                            self.logger.info(f"element_value：{element_value},text:{element_key}")
                            if element_key == "wifi_list":
                                find_result.click()
                                self.logger.info(f"已选择WiFi：{wifi_name}")
                            elif element_key == "wifi_passwd_element":
                                find_result.send_keys(wifi_passwd)
                                self.logger.info(f"已输入密码：{wifi_passwd}")
                            elif element_key == "connect_button":
                                find_result.click()
                                self.logger.info("开始配网...............")
                            else:
                                self.logger.info(f"配网异常")
                        except:
                            if element_key == "wifi_list":
                                self.logger.error(f"未找到wifi：{wifi_name}")
                            elif element_key == "wifi_passwd_element":
                                self.logger.error(f"未找到密码输入框")
                            elif element_key == "connect_button":
                                self.logger.error("未找到连接按钮")
                # 记录开始时间
                start_time = time.time()
                # 等待配网完成
                if "等待配网成功":
                    is_complete = True
                    for device in devices:
                        try:
                            device_complete = WebDriverWait(driver, 60).until(
                                ec.any_of(
                                    ec.text_to_be_present_in_element((By.XPATH,
                                                                      f'//android.widget.TextView[@resource-id="com.ost.pintura:id/tv_name" and '
                                                                      f'@text="{device['id']}"]/following-sibling::*[1]'),
                                                                     "配网成功"),
                                    ec.text_to_be_present_in_element((By.XPATH,
                                                                      f'//android.widget.TextView[@resource-id="com.ost.pintura:id/tv_name" and '
                                                                      f'@text="{device['id']}"]/following-sibling::*[1]'),
                                                                     "连接失败")
                                )
                            )
                        except selenium.common.exceptions.TimeoutException:
                            is_complete = False
                            self.logger.error(f"{device['id']}配网超时")
                            break
                    if is_complete:
                        end_time = time.time()
                        # 总用时
                        total_time = round(end_time - start_time, 2)
                        self.logger.info(f"第{count + 1}次配网已完成，总用时：{total_time}s")
                        count += 1
                if "统计配网结果":
                    # 重新获取元素（配网完成）
                    for device in devices:
                        try:
                            device_complete = WebDriverWait(driver, 10).until(
                                ec.visibility_of_element_located((By.XPATH,
                                                                  f'//android.widget.TextView[@resource-id="com.ost.pintura:id/tv_name" and @text="{device["id"]}"]/following-sibling::*[1]'))
                            )
                            device_result = device_complete.text
                            device['result'] = device_result
                            self.logger.info(f"{device['id']}：{device_result}")
                            # 统计配网成功的次数
                            if device_result == "配网成功":
                                devices_successful += 1
                        except selenium.common.exceptions.NoSuchElementException:
                            self.logger.error(f"获取配网结果元素失败")
                    if devices_successful == len(devices):
                        one_time_setup_successful += 1
                    else:
                        self.logger.info(f"发现配网失败的设备，即将进行截图")
                        self.screenshot_diy(f"第{count}次配网截图.png")
                    self.logger.info(
                        f"配网成功的设备数量：{devices_successful}，失败的设备数量：{len(devices) - devices_successful}")
                    # 处理配网失败的设备
                    device_complete = False
                    for device in devices:
                        if device['result'] == "连接失败":
                            self.logger.info(f"{device['id']}配网失败，尝试重连")
                            try:
                                WebDriverWait(driver, 10).until(
                                    ec.visibility_of_element_located(
                                        (By.XPATH,
                                         f'//android.widget.TextView['
                                         f'@resource-id="com.ost.pintura:id/tv_name" and @text="{device["id"]}"]/following-sibling::*[2]')
                                    )
                                ).click()
                                self.logger.info("点击重连")
                            except selenium.common.exceptions.NoSuchElementException:
                                self.logger.error("找不到配网失败的设备")
                            # 等待重新配网完成
                            try:
                                device_complete = WebDriverWait(driver, 60).until(
                                    ec.any_of(
                                        ec.text_to_be_present_in_element(
                                            (By.XPATH, f'//android.widget.TextView[@resource-id="com.ost.pintura:id/tv_name" and '
                                         f'@text="{device["id"]}"]/following-sibling::*[1]'), "配网成功"),
                                        ec.text_to_be_present_in_element(
                                            (By.XPATH,
                                             f'//android.widget.TextView[@resource-id="com.ost.pintura:id/tv_name" and '
                                             f'@text="{device["id"]}"]/following-sibling::*[1]'), "连接失败"
                                        )
                                        )
                                    )
                            except selenium.common.exceptions.TimeoutException:
                                self.logger.error("二次配网结果元素获取超时")
                            # 重新获取元素并取得文本
                            if device_complete:
                                self.logger.info("二次配网完成，获取元素文本")
                                result_text = WebDriverWait(driver, 10).until(
                                    ec.visibility_of_element_located((By.XPATH,
                                                                      f'//android.widget.TextView[@resource-id="com.ost.pintura:id/tv_name" and @text="{device["id"]}"]/following-sibling::*[1]'))
                                ).text
                                if result_text == "配网成功":
                                    self.logger.info("二次配网成功")
                                    device["second_try_result"] = "二次配网成功"
                                else:
                                    self.logger.info("二次配网失败")
                                    device["second_try_result"] = "二次配网失败"
                                    current_twice_pairing_fail += 1
                    # 生成配网结果文件
                    for device in devices:
                        with open(file_name, "a", encoding=encoding) as f:
                            f.write(f"第{count}次配网：{device["id"]}-{device["result"]}\t{device["second_try_result"]}\t")
                    with open(file_name, "a", encoding=encoding) as f:
                        f.write(f"{total_time}s\n")
                    # 将结果转换未对应的符号,并放入结果列表
                    for device in devices:
                        device_excel_result = self.device_results_report(device["result"], device["second_try_result"])
                        device['pairing_results_excel'].append(device_excel_result)
                    total_time_list.append(total_time)
                    # 返回产品选择界面
                    driver.press_keycode(4)
                    driver.press_keycode(4)
                    driver.press_keycode(4)
                    try:
                        product_selector = WebDriverWait(driver, 10).until(
                            ec.visibility_of_element_located((By.XPATH,

                                                              '//android.webkit.WebView[@text="pages/work/addAhost1[4]"]/android.view.View[3]'))
                        )
                        product_selector.click()
                    except selenium.common.exceptions.TimeoutException:
                        self.logger.error("未找到产品选择按钮")
                if count == circle_times:
                    self.logger.info(f"已测试{circle_times}次，程序即将退出")
                    break
        else:
            self.logger.error(f"{file_name}文件中的配网次数已超过配网次数上限，即将退出程序")
            sys.exit()

        # 统计总的成功率
        fail_count, count, total_succecc_rate_no = self.test_result_statistics(file_name)
        self.logger.info(f"count = {count}，remaining_iterations={remaining_iterations}, circle_times={circle_times}")
        if count > 0:
            total_successful_rate = round((count-fail_count)/count*100, 2)
        else:
            total_successful_rate = 0
        self.logger.info(
            f"本次运行总计配网:{current_iteration}次，成功:{one_time_setup_successful}次，失败"
            f":{current_iteration - one_time_setup_successful}次")
        self.logger.info(f"总的成功率:{total_successful_rate}%")
        # 生成配网结果文件
        with open(file_name, "a", encoding=encoding) as f:
            f.write(f"总的成功率:{total_successful_rate}%\n")
       # 将列表赋值给字典
        pairing_result_dict = {}
        for device in devices:
            pairing_result_dict["device['id']"] = device['pairing_results_excel']
        pairing_result_dict["耗时（S）"] = total_time_list
        # 生成配网结果excel
        pairing_result = pd.DataFrame(pairing_result_dict, columns=column_names)
        pairing_result.index = pairing_result.index + 1
        pairing_result.index.name = "序号"
        self.logger.info(pairing_result)
        # 先读取已有的内容
        if os.path.exists(excel_file_name):
            self.logger.info("检测到文件存在")
            df = pd.read_excel(excel_file_name, index_col=0)
            # 拼接已有的内容到excel
            pairing_result = pd.concat([df, pairing_result], ignore_index=True)
            pairing_result.index = pairing_result.index + 1
        # 添加总计成功率
        total_twice_pairing_fail = current_twice_pairing_fail + twice_paired_fail_excel
        new_index = f"总的成功率：{total_successful_rate}%\t二次配网失败次数：{total_twice_pairing_fail}"
        pairing_result.loc[new_index] = None
        with pd.ExcelWriter(excel_file_name) as writer:
            pairing_result.to_excel(writer, sheet_name="配网结果", index=True)
            self.set_adaptive_column_width(writer, pairing_result, "配网结果")
            # 将总的成功率居左
            worksheet = writer.sheets["配网结果"]
            cell = worksheet[f"A{pairing_result.index.get_loc(new_index) + 2}"]  # 这里的+2是因为python从0开始以及列名一行
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font = Font(bold=True, color="FF0000")
            color_column_names = []
            for device in devices:
                color_column_names.append(device['id'])
            self.set_font_color(writer, pairing_result, color_column_names, "配网结果", color="FFFFFF",
                                colors_condition=True)
            if count != circle_times:
                self.logger.info(f"由于异常原因导致还差{circle_times - count}次配网，程序即将退出")
                sys.exit()
            # 关闭会话防止手机端出错
            driver.quit()


te = bluetooth_pairing_test()
te.detect_network_setup_interface()
