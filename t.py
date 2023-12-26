import os
import pandas as pd
import LoginPage
from Init import get_driver
from VersionSelection import VersionSelection
from LoginPage import LoginPage
import Config
import logging
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import selenium.common.exceptions
from time import sleep
import time
import sys
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font


def excel_reader(file_path):
    try:
        df = pd.read_excel(file_path, index_col=0)
        paired_times = 0
        paired_fail = 0
        twice_paired_fail = 0
        # 获取序列号最后一个数字的值
        for i in df.index[::-1]:
            is_number = isinstance(i, (int, float))
            if is_number:
                paired_times = i
                print(f"最后一次配网次数：{i}")
                break
        # 遍历所有行，统计一次配网失败以及二次配网结果,注意经过excel文件后x变成变得字符×
        for row in df.itertuples():
            if '\u2B55' in row or 'x' in row:
                paired_fail += 1
            if 'x' in row:
                twice_paired_fail += 1
        print(f"paired_times: {paired_times}, paired_fail: {paired_fail}, twice_paired_fail: {twice_paired_fail}")
        return paired_times, paired_fail, twice_paired_fail
    except FileNotFoundError:
        print(f"找不到文件：{file_path}, 请确认是否第一次运行脚本")
        return 0

def excel_remove_rows(file_path, condition):
    is_exist = os.path.exists(file_path)
    print(is_exist)
    df = pd.read_excel(file_path, index_col=0, engine="openpyxl")
    df.index = df.index.astype(str)
    mask = df.index.str.contains(condition)
    df = df[~mask]
    print(df)
    df2 = pd.DataFrame({'Pintura-blt-L000892' : ["x","x", "p"], 'Pintura-blt-Ltest20' : ["x","x", "p"],
                        'Pintura-blt-L000308' : ["x","x", "p"], 'Pintura-blt-L000329' : ["x","x", "p"],
                        '耗时（S）' : [10, 20, 30] })


    print(df2)
    df = pd.concat([df, df2], ignore_index=True)
    df.index = df.index + 1
    print(df)
    with pd.ExcelWriter(file_path) as writer:
        df.to_excel(writer, sheet_name="配网结果", index=True)

def set_adaptive_column_width(writer, data_frame, work_sheet_name="配网结果"):
    # 计算表头的字符宽度
    column_widths = data_frame.columns.to_series().apply(lambda x: len(str(x).encode('utf-8'))).values
    # 计算索引列（序号列）的宽度
    index_width = len(str(data_frame.index.name).encode('utf-8'))
    # 计算每列的最大字符宽度
    max_widths = data_frame.astype(str).map(lambda x: len(x.encode('utf-8'))).max().values
    # 计算整体最大宽度
    widths = np.concatenate(([index_width], column_widths, max_widths))
    # 设置每列的宽度
    worksheet = writer.sheets[work_sheet_name]
    for i, width in enumerate(widths):
        col_letter = get_column_letter(i + 1)
        worksheet.column_dimensions[col_letter].width = width + 2
        # 设置列名（表头）水平和垂直居中
        cell = worksheet[f"{col_letter}1"]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 设置列值（数据）水平和垂直居中
        for row_num in range(2, len(data_frame) + 2):
            cell = worksheet[f"{col_letter}{row_num}"]
            cell.alignment = Alignment(horizontal='center', vertical='center')

def set_font_color(writer, data_frame, column_names, sheet_name="sheet1", color="FFFFFF",
                   colors_condition=False):
    # 设置字体颜色
    font_color = color
    # 获取工作表对象
    worksheet = writer.sheets[sheet_name]
    # 遍历每一列
    for col_to_color in column_names:
        # 获取列的字母表示
        col_letters = get_column_letter(data_frame.columns.get_loc(col_to_color) + 2)
        # 获取列的所有单元格
        column_cells = worksheet[col_letters]
        # 设置字体颜色
        for cell in column_cells:
            if colors_condition and cell.row > 1:
                # 获取cell的值
                if cell.value == "\u2713":
                    cell.font = Font(color="00FF00")
                elif cell.value == "\u2B55":
                    cell.font = Font(color="0000FF")
                elif cell.value == "×":
                    cell.font = Font(color="FF0000")
                else:
                    cell.font = Font(color="000000")
            else:
                if cell.row > 1:  # 排除列名
                    cell.font = Font(color=font_color)


# 调用函数时传递文件路径
# excel_reader("配网结果.xlsx")
# excel_remove_rows("配网结果.xlsx", "总计成功率")

