import pandas as pd
import numpy as np
from pandas import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font


def set_adaptive_column_width(writer, data_frame, work_sheet_name="配网结果"):
    # 计算表头的字符宽度
    column_widths = data_frame.columns.to_series().apply(lambda x: len(str(x).encode('utf-8'))).values
    # 计算索引列（序号列）的宽度
    index_width = len(str(data_frame.index.name).encode('utf-8'))
    # 计算每列的最大字符宽度
    max_widths = data_frame.astype(str).map(lambda x: len(x.encode('utf-8'))).max().values
    # 计算整体最大宽度
    widths = np.concatenate(([index_width], column_widths, max_widths))
    # 设置每列的宽度
    worksheet = writer.sheets[work_sheet_name]
    for i, width in enumerate(widths):
        col_letter = get_column_letter(i + 1)
        worksheet.column_dimensions[col_letter].width = width + 2
        # 设置列名（表头）水平和垂直居中
        cell = worksheet[f"{col_letter}1"]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 设置列值（数据）水平和垂直居中
        for row_num in range(2, len(data_frame) + 2):
            cell = worksheet[f"{col_letter}{row_num}"]
            cell.alignment = Alignment(horizontal='center', vertical='center')


def set_font_color(writer, data_frame, column_names, sheet_name="sheet1", color="FFFFFF", colors_condition=False):
    # 设置字体颜色
    font_color = color
    # 获取工作表对象
    worksheet = writer.sheets[sheet_name]
    # 遍历每一列
    for col_to_color in column_names:
        # 获取列的字母表示
        col_letters = get_column_letter(data_frame.columns.get_loc(col_to_color) + 2)
        # 获取列的所有单元格
        column_cells = worksheet[col_letters]
        # 设置字体颜色
        for cell in column_cells:
            if colors_condition and cell.row > 1:
                # 获取cell的值
                if cell.value == "\u2713":
                    cell.font = Font(color="00FF00")
                elif cell.value == "\u2B55":
                    cell.font = Font(color="0000FF")
                elif cell.value == "×":
                    cell.font = Font(color="FF0000")
                else:
                    cell.font = Font(color="000000")
            else:
                if cell.row > 1:  # 排除列名
                    cell.font = Font(color=font_color)



content = {
    "姓名1111111111111111111111": ["张三1111111111", "李四", "王五", "赵六"],
    "年龄": [18, 192222222222222222, 20, 21],
    "性别": ["\u2713", "\u2B55", "×", "22"],
    "职业": ["IT", "IT", "IT", "IT"]
}
data = pd.DataFrame(content)
data.index = range(1, len(data) + 1)
data.index.name = "序号"
print(data)
# 调用示例：
# 调用示例：
with pd.ExcelWriter("test.xlsx", mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    data.to_excel(writer, sheet_name="配网结果", index=True)
    set_adaptive_column_width(writer, data, work_sheet_name="配网结果")
    color_column_names = ["性别", "职业"]
    set_font_color(writer, data, color_column_names, "配网结果", color="FFFFFF", colors_condition=True)

